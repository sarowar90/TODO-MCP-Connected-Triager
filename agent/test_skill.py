"""Offline checks for the custom shift-handover skill.

Covers the three things that can be wrong with a skill independently of the
model: the SKILL.md frontmatter, whether the skill is discoverable where the
loader looks, and whether its bundled script actually works.

Confirming that the *agent* invokes it needs a key; see README.

Run:
    .venv\\Scripts\\python.exe test_skill.py
"""

import json
import subprocess
import sys
import tempfile
from pathlib import Path

from openpyxl import load_workbook

SKILL_DIR = Path(__file__).resolve().parent / "skills" / "shift-handover"
SKILL_MD = SKILL_DIR / "SKILL.md"
SCRIPT = SKILL_DIR / "scripts" / "build_handover.py"
REFERENCE = SKILL_DIR / "reference" / "format.md"

TICKETS = [
    {
        "ticket_id": "TICK-5001",
        "urgency": "urgent",
        "topic": "technical",
        "team": "engineering",
        "confidence": 0.96,
        "needs_human_review": False,
        "summary": "Platform-wide API 500s since 09:00.",
    },
    {
        "ticket_id": "TICK-5002",
        "urgency": "high",
        "topic": "billing",
        "team": "billing",
        "confidence": 0.91,
        "needs_human_review": False,
        "summary": "Charged twice for this month's subscription.",
    },
    {
        "ticket_id": "TICK-5003",
        "urgency": "high",
        "topic": "account",
        "team": "trust_and_safety",
        "confidence": 0.55,
        "needs_human_review": True,
        "summary": "Unrecognised login from a new country.",
    },
]

passed = failed = 0


def check(name: str, condition: bool, detail: str = "") -> None:
    global passed, failed
    if condition:
        passed += 1
        print(f"  PASS  {name}")
    else:
        failed += 1
        print(f"  FAIL  {name}" + (f" - {detail}" if detail else ""))


def parse_frontmatter(text: str) -> dict[str, str]:
    """Minimal YAML frontmatter reader: enough to validate the block exists,
    is delimited, and carries the fields the loader needs."""
    if not text.startswith("---"):
        return {}
    _, _, rest = text.partition("---\n")
    block, delimiter, _ = rest.partition("\n---")
    if not delimiter:
        return {}

    fields: dict[str, str] = {}
    key = None
    for line in block.splitlines():
        if not line.strip():
            continue
        if not line.startswith((" ", "\t")) and ":" in line:
            key, _, value = line.partition(":")
            key = key.strip()
            fields[key] = value.strip().lstrip(">-").strip()
        elif key:
            fields[key] = (fields[key] + " " + line.strip()).strip()
    return fields


def run_script(args: list[str]) -> subprocess.CompletedProcess:
    return subprocess.run(
        [sys.executable, str(SCRIPT), *args], capture_output=True, text=True
    )


def main() -> int:
    print("skill layout")
    check("SKILL.md exists", SKILL_MD.is_file())
    check("the bundled script exists", SCRIPT.is_file())
    check("the reference resource exists", REFERENCE.is_file())
    check(
        "the script sits in a scripts/ directory the policy trusts",
        SCRIPT.parent.name == "scripts",
    )

    print("\nYAML frontmatter")
    text = SKILL_MD.read_text(encoding="utf-8")
    fields = parse_frontmatter(text)
    check("frontmatter is present and delimited", bool(fields), str(fields)[:80])
    check("name is set", fields.get("name") == "shift-handover")
    check(
        "name matches the directory, which is how it is referenced",
        fields.get("name") == SKILL_DIR.name,
    )
    check("a description is set", bool(fields.get("description")))

    description = fields.get("description", "")
    check(
        "the description says when to use it",
        "after" in description.lower() and "batch" in description.lower(),
        description[:90],
    )
    check(
        "the description says when NOT to use it",
        "do not use" in description.lower(),
        "a description that only says what it does over-triggers",
    )
    check(
        "the description names the artefacts it produces",
        "digest.md" in description and "handover.xlsx" in description,
    )
    check("the description is substantial", len(description) > 200, f"{len(description)} chars")

    print("\nbundled script: happy path")
    with tempfile.TemporaryDirectory() as tmp:
        source = Path(tmp) / "tickets.json"
        destination = Path(tmp) / "handover.xlsx"
        source.write_text(json.dumps(TICKETS), encoding="utf-8")

        result = run_script([str(source), str(destination)])
        check("the script exits 0", result.returncode == 0, result.stderr.strip())
        check("it reports what it wrote", "3 ticket(s)" in result.stdout, result.stdout.strip())
        check("it counts the urgent ticket", "1 urgent" in result.stdout)
        check("it counts the flagged ticket", "1 flagged" in result.stdout)
        check("the workbook exists", destination.is_file())

        book = load_workbook(destination)
        check("it has a Tickets sheet", "Tickets" in book.sheetnames)
        check("it has a Summary sheet", "Summary" in book.sheetnames)

        sheet = book["Tickets"]
        check("one row per ticket plus a header", sheet.max_row == len(TICKETS) + 1)
        check(
            "the header is the agreed column order",
            [c.value for c in sheet[1]]
            == [
                "ticket_id",
                "urgency",
                "topic",
                "team",
                "confidence",
                "needs_human_review",
                "summary",
            ],
        )
        check("the header row is frozen", sheet.freeze_panes == "A2")
        check("an autofilter is set", sheet.auto_filter.ref is not None)

        flagged_row = [r for r in sheet.iter_rows(min_row=2) if r[5].value is True]
        check("the flagged ticket is highlighted", bool(flagged_row) and flagged_row[0][0].fill.fgColor.rgb.endswith("FFF3CD"))

        summary = book["Summary"]
        formulas = [
            c.value
            for row in summary.iter_rows()
            for c in row
            if isinstance(c.value, str) and c.value.startswith("=")
        ]
        check("totals are formulas, not typed numbers", len(formulas) >= 4, str(len(formulas)))
        check("it counts by team", any("D:D" in f for f in formulas))
        check("it counts by urgency", any("B:B" in f for f in formulas))
        check("it counts what needs review", any("F:F" in f for f in formulas))
        check(
            "only Excel-2007-compatible functions are used",
            all("XLOOKUP" not in f and "FILTER" not in f and "UNIQUE" not in f for f in formulas),
        )

    print("\nbundled script: bad input fails loudly")
    with tempfile.TemporaryDirectory() as tmp:
        destination = Path(tmp) / "out.xlsx"

        result = run_script([str(Path(tmp) / "missing.json"), str(destination)])
        check("a missing input exits non-zero", result.returncode != 0)
        check("and says so", "does not exist" in result.stderr)
        check("and writes nothing", not destination.exists())

        bad = Path(tmp) / "bad.json"
        bad.write_text("{not json", encoding="utf-8")
        result = run_script([str(bad), str(destination)])
        check("invalid JSON exits non-zero", result.returncode != 0)
        check("and names the problem", "not valid JSON" in result.stderr)

        bad.write_text("[]", encoding="utf-8")
        result = run_script([str(bad), str(destination)])
        check("an empty batch is refused", result.returncode != 0 and "empty" in result.stderr)

        bad.write_text(json.dumps([{"ticket_id": "TICK-1"}]), encoding="utf-8")
        result = run_script([str(bad), str(destination)])
        check("a missing field is refused", result.returncode != 0)
        check(
            "and the message names the offending ticket and fields",
            "TICK-1" in result.stderr and "urgency" in result.stderr,
            result.stderr.strip(),
        )

        result = run_script([str(bad)])
        check("wrong argument count exits non-zero", result.returncode != 0)

    print(f"\n{passed} passed, {failed} failed")
    return 1 if failed else 0


if __name__ == "__main__":
    raise SystemExit(main())
