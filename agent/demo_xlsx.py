"""Demonstration: the handover workbook opens and is correct.

Runs offline — no API key. It builds a workbook matching the spec the digest
step asks the xlsx skill to produce, then validates it with the *same*
`check_workbook()` the agent's goal check uses, and re-opens it to confirm the
formulas and values are real.

What this proves and what it does not
------------------------------------
Proves: the target format is well-formed, openpyxl can open it, the goal check
accepts a correct workbook and rejects broken ones, and the toolchain in
requirements.txt can produce a real .xlsx.

Does not prove: that the pre-built xlsx skill, driven by the model, produces
this. That needs a key and a populated skills/ directory — see README.

    .venv\\Scripts\\python.exe demo_xlsx.py
"""

import sys

from openpyxl import Workbook, load_workbook

from fs_policy import OUTBOX, ensure_workspace
from loop import (
    WORKBOOK_HEADERS,
    WORKBOOK_NAME,
    WORKBOOK_SHEET,
    check_workbook,
)

TICKETS = [
    ("TICK-5001", "urgent", "technical", "engineering", 0.96, False,
     "Platform-wide API 500s since 09:00 blocking the whole team."),
    ("TICK-5002", "high", "billing", "billing", 0.91, False,
     "Charged twice for this month's subscription; second payment failed."),
    ("TICK-5003", "high", "account", "trust_and_safety", 0.88, True,
     "Unrecognised login from a new country; possible compromise."),
]


def build(path) -> None:
    book = Workbook()
    sheet = book.active
    sheet.title = WORKBOOK_SHEET
    sheet.append(list(WORKBOOK_HEADERS))
    for row in TICKETS:
        sheet.append(list(row))

    for column, width in zip("ABCDEFG", (12, 10, 16, 18, 12, 20, 60)):
        sheet.column_dimensions[column].width = width

    # Counts by formula, not typed totals, so the figures recalculate when a
    # row is edited. COUNTIF is Excel-2007 and evaluates in LibreOffice too.
    summary = book.create_sheet("Summary")
    summary.append(["team", "tickets"])
    for index, team in enumerate(sorted({t[3] for t in TICKETS}), start=2):
        summary.cell(row=index, column=1, value=team)
        summary.cell(
            row=index,
            column=2,
            value=f"=COUNTIF({WORKBOOK_SHEET}!D:D,A{index})",
        )

    start = summary.max_row + 2
    summary.cell(row=start, column=1, value="urgency")
    summary.cell(row=start, column=2, value="tickets")
    for offset, urgency in enumerate(sorted({t[1] for t in TICKETS}), start=1):
        summary.cell(row=start + offset, column=1, value=urgency)
        summary.cell(
            row=start + offset,
            column=2,
            value=f"=COUNTIF({WORKBOOK_SHEET}!B:B,A{start + offset})",
        )

    book.save(path)


def main() -> int:
    ensure_workspace()
    path = OUTBOX / WORKBOOK_NAME
    ticket_ids = [t[0] for t in TICKETS]

    print("=" * 68)
    print("1. BUILD THE WORKBOOK")
    print("=" * 68)
    build(path)
    print(f"  wrote {path.name} ({path.stat().st_size:,} bytes)")

    print("\n" + "=" * 68)
    print("2. VALIDATE WITH THE AGENT'S OWN GOAL CHECK")
    print("=" * 68)
    met, reason = check_workbook(ticket_ids)
    print(f"  check_workbook -> {'OK' if met else 'FAILED: ' + reason}")

    print("\n" + "=" * 68)
    print("3. RE-OPEN AND READ IT BACK")
    print("=" * 68)
    book = load_workbook(path)
    print(f"  sheets       {book.sheetnames}")
    sheet = book[WORKBOOK_SHEET]
    print(f"  dimensions   {sheet.max_row} rows x {sheet.max_column} cols")
    print(f"  header       {[c.value for c in sheet[1]]}")
    for row in sheet.iter_rows(min_row=2, max_col=4, values_only=True):
        print(f"  row          {row}")

    summary = book["Summary"]
    formulas = [
        c.value
        for row in summary.iter_rows()
        for c in row
        if isinstance(c.value, str) and c.value.startswith("=")
    ]
    print(f"  formulas     {len(formulas)} (e.g. {formulas[0] if formulas else 'none'})")

    print("\n" + "=" * 68)
    print("4. VERIFY")
    print("=" * 68)
    checks = [
        ("the goal check accepts it", met),
        ("it opens as a real workbook", WORKBOOK_SHEET in book.sheetnames),
        ("the header matches the agreed columns",
         [str(c.value) for c in sheet[1]] == list(WORKBOOK_HEADERS)),
        ("one row per ticket", sheet.max_row == len(TICKETS) + 1),
        ("every ticket id is present",
         all(any(tid == c.value for row in sheet.iter_rows() for c in row)
             for tid in ticket_ids)),
        ("the flagged ticket is marked for review",
         any(row[5] is True for row in sheet.iter_rows(min_row=2, values_only=True))),
        ("counts are formulas, not typed totals", len(formulas) >= 2),
        ("a summary sheet exists", "Summary" in book.sheetnames),
    ]

    # The check must also reject a workbook that is wrong, or it proves nothing.
    broken = OUTBOX / "_broken.xlsx"
    bad = Workbook()
    bad.active.title = "Wrong"
    bad.save(broken)
    rejected, why = check_workbook(ticket_ids)  # still points at the good file
    bad_path_ok = True
    try:
        path.rename(OUTBOX / "_stash.xlsx")
        broken.rename(path)
        rejected, why = check_workbook(ticket_ids)
        checks.append(("a workbook without the Tickets sheet is rejected", not rejected))
    finally:
        path.unlink(missing_ok=True)
        (OUTBOX / "_stash.xlsx").rename(path)
        broken.unlink(missing_ok=True)

    print()
    ok = True
    for name, condition in checks:
        print(f"  {'PASS' if condition else 'FAIL'}  {name}")
        ok = ok and bool(condition)

    print("\n" + "=" * 68)
    print(f"WORKBOOK VALID  ({path})" if ok else "WORKBOOK INVALID")
    print("=" * 68)
    print("\nOpen it to confirm by eye:")
    print(f"  start {path}")
    return 0 if ok and bad_path_ok else 1


if __name__ == "__main__":
    raise SystemExit(main())
