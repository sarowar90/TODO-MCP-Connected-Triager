#!/usr/bin/env python3
"""Build the shift-handover workbook from a tickets JSON file.

Bundled resource of the shift-handover skill. The agent writes tickets.json and
runs this rather than hand-writing openpyxl, so the column order, formula style
and formatting are fixed rather than re-invented per run.

    python build_handover.py <tickets.json> <handover.xlsx>

Exits non-zero with a specific message on bad input, so a failure tells the
agent what to fix instead of leaving a half-written file behind.
"""

import json
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:  # pragma: no cover
    print("error: openpyxl is not installed", file=sys.stderr)
    raise SystemExit(2)

HEADERS = (
    "ticket_id",
    "urgency",
    "topic",
    "team",
    "confidence",
    "needs_human_review",
    "summary",
)
SHEET = "Tickets"
SUMMARY = "Summary"
WIDTHS = (14, 10, 16, 18, 12, 20, 70)

# Urgency order for the summary, worst first — the next shift reads top-down.
URGENCY_ORDER = ("urgent", "high", "normal", "low")

HEADER_FILL = PatternFill("solid", fgColor="DDDDDD")
FLAG_FILL = PatternFill("solid", fgColor="FFF3CD")


def fail(message: str) -> None:
    print(f"error: {message}", file=sys.stderr)
    raise SystemExit(1)


def load(path: Path) -> list[dict]:
    if not path.is_file():
        fail(f"{path} does not exist")
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        fail(f"{path.name} is not valid JSON: {exc}")

    if not isinstance(data, list):
        fail(f"{path.name} must contain a JSON array of ticket objects")
    if not data:
        fail(f"{path.name} is empty; there is nothing to hand over")

    for index, row in enumerate(data):
        if not isinstance(row, dict):
            fail(f"entry {index} is not an object")
        missing = [key for key in HEADERS if key not in row]
        if missing:
            fail(f"entry {index} ({row.get('ticket_id', '?')}) is missing: {', '.join(missing)}")
    return data


def build(tickets: list[dict], destination: Path) -> None:
    book = Workbook()
    sheet = book.active
    sheet.title = SHEET

    sheet.append(list(HEADERS))
    for cell in sheet[1]:
        cell.font = Font(name="Arial", bold=True)
        cell.fill = HEADER_FILL

    for ticket in tickets:
        sheet.append([ticket[key] for key in HEADERS])

    for column, width in zip("ABCDEFG", WIDTHS):
        sheet.column_dimensions[column].width = width
    for row in sheet.iter_rows(min_row=2):
        row[6].alignment = Alignment(wrap_text=True, vertical="top")
        # Highlight the rows a human still has to look at.
        if row[5].value is True:
            for cell in row:
                cell.fill = FLAG_FILL

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:G{sheet.max_row}"

    summary = book.create_sheet(SUMMARY)
    summary["A1"] = "team"
    summary["B1"] = "tickets"
    for cell in (summary["A1"], summary["B1"]):
        cell.font = Font(name="Arial", bold=True)

    teams = sorted({str(t["team"]) for t in tickets})
    for offset, team in enumerate(teams, start=2):
        summary.cell(row=offset, column=1, value=team)
        # A formula, not a count computed here: the sheet must stay correct if
        # someone edits a row on the Tickets sheet.
        summary.cell(row=offset, column=2, value=f"=COUNTIF({SHEET}!D:D,A{offset})")

    start = len(teams) + 3
    summary.cell(row=start, column=1, value="urgency").font = Font(name="Arial", bold=True)
    summary.cell(row=start, column=2, value="tickets").font = Font(name="Arial", bold=True)

    present = [u for u in URGENCY_ORDER if any(str(t["urgency"]) == u for t in tickets)]
    for offset, urgency in enumerate(present, start=1):
        row = start + offset
        summary.cell(row=row, column=1, value=urgency)
        summary.cell(row=row, column=2, value=f"=COUNTIF({SHEET}!B:B,A{row})")

    flagged = start + len(present) + 2
    summary.cell(row=flagged, column=1, value="needs human review").font = Font(
        name="Arial", bold=True
    )
    summary.cell(row=flagged, column=2, value=f"=COUNTIF({SHEET}!F:F,TRUE)")

    summary.column_dimensions["A"].width = 22
    summary.column_dimensions["B"].width = 10

    destination.parent.mkdir(parents=True, exist_ok=True)
    book.save(destination)


def main(argv: list[str]) -> int:
    if len(argv) != 2:
        print(__doc__.strip().splitlines()[3].strip(), file=sys.stderr)
        return 2

    source, destination = Path(argv[0]), Path(argv[1])
    tickets = load(source)
    build(tickets, destination)

    flagged = sum(1 for t in tickets if t.get("needs_human_review") is True)
    urgent = sum(1 for t in tickets if str(t.get("urgency")) == "urgent")
    print(
        f"wrote {destination.name}: {len(tickets)} ticket(s), "
        f"{urgent} urgent, {flagged} flagged for review"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
