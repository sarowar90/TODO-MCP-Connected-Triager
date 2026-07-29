"""Offline checks for multi-step execution.

Covers the parts of step 5 that don't need the model: that the goal is
decomposed into an ordered plan before work starts, that step status tracks
execution, and — most importantly — that the digest step's goal check actually
holds it to accounting for every ticket the earlier steps produced.

Run:
    .venv\\Scripts\\python.exe test_plan.py
"""

import asyncio

from fs_policy import OUTBOX, ensure_workspace
from loop import DIGEST_NAME, LoopOutcome, StepResult, make_digest_goal, ticket_goal
from plan import Plan, build_plan, load_inbox
from triage_tools import TriageSession

passed = failed = 0


def check(name: str, condition: bool, detail: str = "") -> None:
    global passed, failed
    if condition:
        passed += 1
        print(f"  PASS  {name}")
    else:
        failed += 1
        print(f"  FAIL  {name}" + (f" - {detail}" if detail else ""))


def _step(name: str, met: bool, ticket: dict | None = None) -> StepResult:
    session = TriageSession()
    session.ticket = ticket
    return StepResult(name, met, session, 1, 3, 0.01, "success")


async def main() -> int:
    ensure_workspace()

    print("plan decomposition")
    messages = load_inbox()
    check("inbox messages are discovered", len(messages) >= 2, f"found {len(messages)}")

    plan = build_plan(messages)
    check(
        "plan has one step per message plus a digest step",
        len(plan.steps) == len(messages) + 1,
        f"{len(plan.steps)} steps for {len(messages)} messages",
    )
    check("the digest is sequenced last", plan.steps[-1].key == "digest")
    check(
        "triage steps come first, in inbox order",
        [s.key for s in plan.steps[:-1]] == [f"triage:{m[0]}" for m in messages],
    )
    check("every step starts pending", all(s.status == "pending" for s in plan.steps))

    print("\nstatus tracking")
    first = plan.steps[0].key
    plan.set(first, "running")
    check("a step can be marked running", plan.steps[0].status == "running")
    plan.set(first, "done", "-> TICK-5001")
    check("a step can be marked done", plan.steps[0].status == "done")
    check("a note is recorded alongside", "TICK-5001" in plan.steps[0].note)
    plan.set("digest", "skipped", "no tickets were filed")
    check("the digest can be skipped", plan.steps[-1].status == "skipped")

    print("\ntriage goal")
    empty = TriageSession()
    met, reason = ticket_goal(empty)
    check("unmet before create_ticket", not met and "stopped before" in reason)

    empty.rejections.append("team 'sales' is not valid")
    met, reason = ticket_goal(empty)
    check("a rejection is surfaced as the reason", "rejected" in reason)

    filed = TriageSession()
    filed.ticket = {"ticket_id": "TICK-9999"}
    met, reason = ticket_goal(filed)
    check(
        "filing without writing the file leaves the goal unmet",
        not met and "never wrote" in reason,
    )

    print("\ndigest goal")
    digest = OUTBOX / DIGEST_NAME
    stale = digest.read_text(encoding="utf-8") if digest.is_file() else None
    try:
        digest.unlink(missing_ok=True)
        goal = make_digest_goal(["TICK-5001", "TICK-5002"])

        met, reason = goal(TriageSession())
        check("unmet when the digest is absent", not met and "does not exist" in reason)

        digest.write_text("# Handover\n\nTICK-5001 was urgent.\n", encoding="utf-8")
        met, reason = goal(TriageSession())
        check(
            "unmet when a ticket is unaccounted for",
            not met and "TICK-5002" in reason,
            reason,
        )

        digest.write_text(
            "# Handover\n\n| id | urgency |\n| TICK-5001 | urgent |\n"
            "| TICK-5002 | high |\n",
            encoding="utf-8",
        )
        met, _ = goal(TriageSession())
        check("met once every ticket is accounted for", met)

        check(
            "an empty batch needs no ticket ids",
            make_digest_goal([])(TriageSession())[0],
        )
    finally:
        digest.unlink(missing_ok=True)
        if stale is not None:
            digest.write_text(stale, encoding="utf-8")

    print("\nworkbook deliverable")
    from loop import WORKBOOK_HEADERS, WORKBOOK_NAME, WORKBOOK_SHEET, check_workbook

    workbook = OUTBOX / WORKBOOK_NAME
    stash = workbook.read_bytes() if workbook.is_file() else None
    try:
        workbook.unlink(missing_ok=True)
        met, reason = check_workbook(["TICK-1"])
        check("unmet when the workbook is absent", not met and "does not exist" in reason)

        workbook.write_bytes(b"this is not a spreadsheet")
        met, reason = check_workbook(["TICK-1"])
        check("a non-xlsx file is rejected, not merely counted", not met and "does not open" in reason)

        from openpyxl import Workbook

        book = Workbook()
        book.active.title = WORKBOOK_SHEET
        book.active.append(list(WORKBOOK_HEADERS))
        book.save(workbook)
        met, reason = check_workbook(["TICK-1"])
        check("a workbook with no row for a ticket is rejected", not met and "no row" in reason)

        book = Workbook()
        book.active.title = WORKBOOK_SHEET
        book.active.append(list(WORKBOOK_HEADERS)[:-1])  # drop a column
        book.active.append(["TICK-1"])
        book.save(workbook)
        met, reason = check_workbook(["TICK-1"])
        check("a missing column is rejected", not met and "missing column" in reason)

        book = Workbook()
        book.active.title = WORKBOOK_SHEET
        book.active.append(list(WORKBOOK_HEADERS))
        book.active.append(["TICK-1", "high", "billing", "billing", 0.9, False, "x"])
        book.save(workbook)
        met, _ = check_workbook(["TICK-1"])
        check("a correct workbook is accepted", met)
    finally:
        workbook.unlink(missing_ok=True)
        if stash is not None:
            workbook.write_bytes(stash)

    print("\noutcome aggregation")
    outcome = LoopOutcome()
    outcome.steps = [
        _step("triage a", True, {"ticket_id": "TICK-1", "urgency": "high"}),
        _step("triage b", True, {"ticket_id": "TICK-2", "urgency": "low"}),
        _step("write digest", True),
    ]
    check("goal met when every step succeeded", outcome.goal_met)
    check("turns are summed across steps", outcome.turns == 9)
    check("tickets are collected from the steps", len(outcome.tickets) == 2)

    outcome.steps[1] = _step("triage b", False)
    check("one failed step fails the run", not outcome.goal_met)
    check(
        "a partial batch still reports the tickets it did file",
        len(outcome.tickets) == 1,
    )

    check("an empty run is not counted as met", not LoopOutcome().goal_met)

    print(f"\n{passed} passed, {failed} failed")
    return 1 if failed else 0


if __name__ == "__main__":
    raise SystemExit(asyncio.run(main()))
