"""Demonstration: the shift-handover skill's workflow, end to end.

Runs offline — no API key. It walks the exact steps SKILL.md instructs the
agent to take, using the real bundled script and the real permission policy,
and checks the result with the agent's own goal check.

What this proves and what it does not
-------------------------------------
Proves: the skill is discoverable where the loader looks, its instructions
describe a workflow that actually works, the command it tells the agent to run
is permitted by the policy (and the obvious variations are not), and the
artefacts it produces pass the goal check.

Does not prove: that the model reads the description and chooses to invoke the
skill. That needs a key — see README.

    .venv\\Scripts\\python.exe demo_skill.py
"""

import json
import subprocess
import sys

from fs_policy import OUTBOX, ensure_workspace
from loop import CUSTOM_SKILL, SKILLS_DIR, WORKBOOK_NAME, check_workbook, skills_available
from permissions import Tier, classify

SCRIPT = SKILLS_DIR / CUSTOM_SKILL / "scripts" / "build_handover.py"

TICKETS = [
    {
        "ticket_id": "TICK-5001", "urgency": "urgent", "topic": "technical",
        "team": "engineering", "confidence": 0.96, "needs_human_review": False,
        "summary": "Platform-wide API 500s since 09:00 blocking the whole team.",
    },
    {
        "ticket_id": "TICK-5002", "urgency": "high", "topic": "billing",
        "team": "billing", "confidence": 0.91, "needs_human_review": False,
        "summary": "Charged twice for this month's subscription.",
    },
    {
        "ticket_id": "TICK-5003", "urgency": "high", "topic": "account",
        "team": "trust_and_safety", "confidence": 0.55, "needs_human_review": True,
        "summary": "Unrecognised login from a new country; possible compromise.",
    },
]


def rule(title: str) -> None:
    print("\n" + "=" * 70)
    print(title)
    print("=" * 70)


def main() -> int:
    ensure_workspace()
    checks: list[tuple[str, bool]] = []

    rule("1. THE SKILL IS DISCOVERABLE")
    skill_md = SKILLS_DIR / CUSTOM_SKILL / "SKILL.md"
    print(f"  {skill_md.relative_to(SKILLS_DIR.parent)}")
    print(f"  loader sees skills: {skills_available()}")
    checks.append(("the skill is on disk where the loader looks", skill_md.is_file()))
    checks.append(("the loader reports skills available", skills_available()))

    rule("2. THE COMMAND SKILL.md TELLS THE AGENT TO RUN IS PERMITTED")
    tickets_json = OUTBOX / "tickets.json"
    workbook = OUTBOX / WORKBOOK_NAME
    command = f"python {SCRIPT} {tickets_json} {workbook}"
    decision = classify("Bash", {"command": command})
    print(f"  {command}")
    print(f"  -> {decision.tier.value}: {decision.reason}")
    checks.append(("the documented command is allowed", decision.tier is Tier.AUTO))

    # The policy must still refuse the obvious ways to abuse it.
    for label, variant in [
        ("output redirected outside the outbox", f"python {SCRIPT} {tickets_json} /app/lib/x.xlsx"),
        ("a second command appended", f"python {SCRIPT} {tickets_json} {workbook}; rm -rf /"),
        ("a different interpreter", f"sh {SCRIPT}"),
    ]:
        tier = classify("Bash", {"command": variant}).tier
        print(f"  refused: {label} -> {tier.value}")
        checks.append((f"refused: {label}", tier is Tier.DENY))

    rule("3. STEP 2 OF THE SKILL: WRITE tickets.json")
    tickets_json.write_text(json.dumps(TICKETS, indent=2), encoding="utf-8")
    print(f"  wrote {tickets_json.name} ({len(TICKETS)} tickets)")

    rule("4. STEP 3 OF THE SKILL: RUN THE BUNDLED SCRIPT")
    result = subprocess.run(
        [sys.executable, str(SCRIPT), str(tickets_json), str(workbook)],
        capture_output=True, text=True,
    )
    print(f"  exit {result.returncode}")
    for line in (result.stdout + result.stderr).strip().splitlines():
        print(f"  {line}")
    checks.append(("the bundled script succeeds", result.returncode == 0))

    rule("5. THE RESULT PASSES THE AGENT'S GOAL CHECK")
    met, reason = check_workbook([t["ticket_id"] for t in TICKETS])
    print(f"  check_workbook -> {'OK' if met else 'FAILED: ' + reason}")
    checks.append(("the workbook satisfies the goal check", met))

    from openpyxl import load_workbook

    book = load_workbook(workbook)
    sheet = book["Tickets"]
    print(f"  sheets: {book.sheetnames}, rows: {sheet.max_row}")
    flagged = [r[0].value for r in sheet.iter_rows(min_row=2) if r[5].value is True]
    print(f"  flagged for review: {flagged}")
    checks.append(("the flagged ticket is carried through", flagged == ["TICK-5003"]))

    rule("VERIFY")
    ok = True
    for name, condition in checks:
        print(f"  {'PASS' if condition else 'FAIL'}  {name}")
        ok = ok and condition

    print("\n" + "=" * 70)
    print("SKILL WORKFLOW VALID" if ok else "SKILL WORKFLOW BROKEN")
    print("=" * 70)
    print(f"\n  open the workbook: start {workbook}")
    print("\n  Not covered here: whether the model reads the description and")
    print("  chooses to invoke the skill. That needs an API key.")

    tickets_json.unlink(missing_ok=True)
    return 0 if ok else 1


if __name__ == "__main__":
    raise SystemExit(main())
